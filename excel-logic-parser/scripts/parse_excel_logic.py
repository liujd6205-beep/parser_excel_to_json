# -*- coding: utf-8 -*-
"""
Excel 质控逻辑文档解析器
========================
将包含5列固定表头的Excel逻辑文档转换为结构化JSON，供后续AI编写Python脚本使用。

固定表头（按列顺序）：
  A: 字段名称        B: 字段逻辑场景
  C: 质检结果提示文案  D: 质控解决方案
  E: 调整日志

用法:
  python parse_excel_logic.py <excel文件路径> [输出JSON路径]

  若不指定输出路径，则输出到同目录下，文件名与Excel相同但扩展名为.json

依赖:
  pip install openpyxl

修改指南:
  - 修改 COLUMN_MAP 可调整列映射
  - 修改 ACTION_PATTERNS 可调整调整日志的解析规则
  - 修改 SOLUTION_PATTERNS 可调整质控解决方案的解析规则
  - 修改 RED_COLOR_PATTERNS 可调整红色字体检测规则
"""

import sys
import os
import re
import json
import argparse

try:
    import openpyxl
except ImportError:
    print("错误: 需要安装 openpyxl。请运行: pip install openpyxl")
    sys.exit(1)


# ============================================================
#  配置区：修改这里即可适配不同的Excel格式
# ============================================================

# 列映射：表头名称 -> 列索引（从1开始）
COLUMN_MAP = {
    "field_name":       1,   # A列: 字段名称
    "logic_scene":      2,   # B列: 字段逻辑场景
    "prompt_text":      3,   # C列: 质检结果提示文案
    "solution":         4,   # D列: 质控解决方案
    "adjust_log":       5,   # E列: 调整日志
}

# 数据起始行（跳过表头行）
DATA_START_ROW = 2

# 数据结束行（None表示自动检测到最后一行有数据的行）
DATA_END_ROW = None

# 工作表名称（None表示使用第一个工作表）
SHEET_NAME = None

# 红色字体颜色模式（匹配这些RGB值的字体视为红色，需特别关注）
# 格式：ARGB字符串，匹配包含这些模式即视为红色
RED_COLOR_PATTERNS = ["FF0000"]

# 调整日志解析规则
# 格式：(正则模式, action_type, 是否标记为红色)
# 按顺序匹配，第一个匹配到的规则生效
ACTION_PATTERNS = [
    (r"^删除逻辑",            "delete_logic",            True),   # 删除逻辑（红色字体标记）
    (r"^新增逻辑",            "add_logic",               False),  # 新增逻辑
    (r"^删除访视方式为面对面",  "delete_visit_face_to_face", False), # 删除访视方式为面对面
    (r"^逻辑调整",            "adjust_logic",            False),  # 逻辑调整（范围值调整等）
]

# 质控解决方案解析规则
# 返回: {"type": ..., "ideal_data": ..., "description": ...}
# 按顺序匹配，第一个匹配到的规则生效
def parse_solution(text):
    """解析质控解决方案文本，返回结构化的解决方案信息"""
    if not text:
        return {"type": "unknown", "ideal_data": None, "description": ""}

    t = str(text).strip()

    # 1. 只提示不修改 -> ideal_data = None
    if t == "只提示不修改":
        return {"type": "prompt_only", "ideal_data": None, "description": "仅提示不修改"}

    # 2. 清空 -> ideal_data = "清空"（注意：是这两个字本身，不是空字符串）
    if t == "清空":
        return {"type": "modify_clear", "ideal_data": "清空", "description": "清空该字段值"}

    # 3. 选xxx -> 选择固定值
    m = re.match(r"^选(.+)$", t)
    if m:
        return {"type": "modify_select", "ideal_data": m.group(1), "description": f"选{m.group(1)}"}

    # 4. 添加勾选"xxx" 或 添加勾选xxx -> 勾选项（必须在"添加xxx"之前匹配）
    m = re.match(r"^添加勾选(.+)$", t)
    if m:
        val = m.group(1).strip()
        # 去除首尾引号（支持中文引号\u201c\u201d、英文引号"等）
        if len(val) >= 2 and val[0] in '\u201c\u201d"' and val[-1] in '\u201c\u201d"':
            val = val[1:-1]
        return {"type": "modify_add_check", "ideal_data": val, "description": f"勾选{val}"}

    # 5. 添加xxx -> 添加内容
    m = re.match(r"^添加(.+)$", t)
    if m:
        return {"type": "modify_add", "ideal_data": m.group(1), "description": f"添加{m.group(1)}"}

    # 6. 必须包含xxx -> 必须包含某内容
    m = re.match(r"^必须包含(.+)$", t)
    if m:
        return {"type": "modify_contain", "ideal_data": m.group(1), "description": f"必须包含{m.group(1)}"}

    # 7. xxx随机取值 -> 范围内随机取值
    m = re.match(r"^(.+?)随机取值$", t)
    if m:
        range_str = m.group(1).strip()
        # 解析范围，如 "9-12", "36-37.2次", "1-7次", "4-5.9", "2.5-3"
        rm = re.match(r"(\d+\.?\d*)\s*[-~]\s*(\d+\.?\d*)", range_str)
        if rm:
            min_val = float(rm.group(1))
            max_val = float(rm.group(2))
            return {
                "type": "modify_random",
                "ideal_data": {"min": min_val, "max": max_val},
                "description": f"在[{min_val}, {max_val}]范围内随机取值"
            }
        return {"type": "modify_random", "ideal_data": range_str, "description": f"随机取值: {range_str}"}

    # 8. 填xxx -> 填写固定文本
    m = re.match(r"^填(.+)$", t)
    if m:
        return {"type": "modify_fill", "ideal_data": m.group(1), "description": f"填写: {m.group(1)}"}

    # 9. 选择异常 -> 选择异常选项
    if t == "选择异常":
        return {"type": "modify_select", "ideal_data": "异常", "description": "选择异常"}

    # 10. 默认：直接填写该文本
    return {"type": "modify_fill", "ideal_data": t, "description": f"填写: {t}"}


# ============================================================
#  核心解析逻辑（一般不需要修改以下内容）
# ============================================================

def is_red_font(cell):
    """检测单元格是否包含红色字体"""
    if cell.font and cell.font.color:
        c = cell.font.color
        if c.type == 'rgb' and c.rgb:
            rgb_str = str(c.rgb).upper()
            for pattern in RED_COLOR_PATTERNS:
                if pattern in rgb_str and rgb_str != 'FF000000':
                    return True
    return False


def parse_action(adjust_log, cell):
    """解析调整日志，返回action信息"""
    if not adjust_log:
        return {"type": "none", "detail": "", "is_red": False}

    text = str(adjust_log).strip()

    for pattern, action_type, default_red in ACTION_PATTERNS:
        if re.match(pattern, text):
            return {
                "type": action_type,
                "detail": text,
                "is_red": default_red or is_red_font(cell)
            }

    return {"type": "unknown", "detail": text, "is_red": is_red_font(cell)}


def parse_excel_to_json(excel_path, output_path=None, sheet_name=None,
                        data_start_row=None, data_end_row=None):
    """
    解析Excel逻辑文档为结构化JSON

    Args:
        excel_path: Excel文件路径
        output_path: 输出JSON文件路径（None则自动生成）
        sheet_name: 工作表名称（None使用第一个）
        data_start_row: 数据起始行（None使用配置默认值）
        data_end_row: 数据结束行（None自动检测）

    Returns:
        dict: 解析后的结构化数据
    """
    # 加载配置
    if data_start_row is None:
        data_start_row = DATA_START_ROW
    if sheet_name is None:
        sheet_name = SHEET_NAME

    # 加载Excel
    wb = openpyxl.load_workbook(excel_path, rich_text=True)

    if sheet_name:
        ws = wb[sheet_name]
    else:
        ws = wb.worksheets[0]

    # 自动检测结束行
    if data_end_row is None:
        data_end_row = ws.max_row

    # 解析数据
    entries = []
    current_field_name = None
    current_field_note = None

    for row in range(data_start_row, data_end_row + 1):
        a_val = ws.cell(row=row, column=COLUMN_MAP["field_name"]).value
        b_val = ws.cell(row=row, column=COLUMN_MAP["logic_scene"]).value
        c_val = ws.cell(row=row, column=COLUMN_MAP["prompt_text"]).value
        d_val = ws.cell(row=row, column=COLUMN_MAP["solution"]).value
        e_val = ws.cell(row=row, column=COLUMN_MAP["adjust_log"]).value

        # 跳过空行
        if b_val is None and c_val is None and d_val is None and e_val is None:
            continue

        # 更新当前字段名（A列有值时更新，支持合并单元格的继承）
        if a_val is not None:
            field_text = str(a_val).strip()
            if '\n' in field_text:
                parts = field_text.split('\n', 1)
                current_field_name = parts[0].strip()
                current_field_note = parts[1].strip()
            else:
                current_field_name = field_text
                current_field_note = None

        # 解析action
        e_cell = ws.cell(row=row, column=COLUMN_MAP["adjust_log"])
        action = parse_action(e_val, e_cell)

        # 解析solution
        solution = parse_solution(d_val)

        # 检测B列是否红色字体
        b_cell = ws.cell(row=row, column=COLUMN_MAP["logic_scene"])
        b_is_red = is_red_font(b_cell)

        entry = {
            "row": row,
            "field_name": current_field_name,
            "field_note": current_field_note,
            "logic_scene": str(b_val).strip() if b_val else None,
            "prompt_text": str(c_val).strip() if c_val else None,
            "solution": solution,
            "action": action,
            "logic_scene_is_red": b_is_red,
        }

        entries.append(entry)

    # 按字段名分组
    grouped = {}
    for entry in entries:
        fn = entry["field_name"]
        if fn not in grouped:
            grouped[fn] = {
                "field_name": fn,
                "field_note": entry["field_note"],
                "logics": []
            }
        grouped[fn]["logics"].append(entry)

    result_fields = list(grouped.values())

    # 统计信息
    action_stats = {}
    solution_stats = {}
    for e in entries:
        at = e["action"]["type"]
        action_stats[at] = action_stats.get(at, 0) + 1
        st = e["solution"]["type"]
        solution_stats[st] = solution_stats.get(st, 0) + 1

    summary = {
        "total_fields": len(result_fields),
        "total_logics": len(entries),
        "action_stats": action_stats,
        "solution_stats": solution_stats,
        "red_font_logics": sum(1 for e in entries if e["action"]["is_red"]),
        "red_scene_logics": sum(1 for e in entries if e["logic_scene_is_red"]),
    }

    output = {
        "summary": summary,
        "fields": result_fields
    }

    # 写入JSON
    if output_path is None:
        base, _ = os.path.splitext(excel_path)
        output_path = base + ".json"

    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump(output, f, ensure_ascii=False, indent=2)

    # 打印统计信息
    print(f"解析完成: {output_path}")
    print(f"  字段总数: {summary['total_fields']}")
    print(f"  逻辑条目总数: {summary['total_logics']}")
    print(f"  红色字体(调整日志): {summary['red_font_logics']}")
    print(f"  红色字体(逻辑场景): {summary['red_scene_logics']}")
    print(f"\n  Action类型统计:")
    for k, v in sorted(action_stats.items(), key=lambda x: -x[1]):
        print(f"    {k}: {v}")
    print(f"\n  Solution类型统计:")
    for k, v in sorted(solution_stats.items(), key=lambda x: -x[1]):
        print(f"    {k}: {v}")

    return output


# ============================================================
#  命令行入口
# ============================================================

def main():
    parser = argparse.ArgumentParser(
        description="Excel质控逻辑文档解析器 - 将Excel转换为结构化JSON"
    )
    parser.add_argument("excel_path", help="Excel文件路径")
    parser.add_argument("output_path", nargs="?", default=None,
                        help="输出JSON路径（默认与Excel同目录）")
    parser.add_argument("--sheet", default=None, help="工作表名称（默认第一个）")
    parser.add_argument("--start-row", type=int, default=None,
                        help=f"数据起始行（默认{DATA_START_ROW}）")
    parser.add_argument("--end-row", type=int, default=None,
                        help="数据结束行（默认自动检测）")

    args = parser.parse_args()

    if not os.path.exists(args.excel_path):
        print(f"错误: 文件不存在: {args.excel_path}")
        sys.exit(1)

    result = parse_excel_to_json(
        excel_path=args.excel_path,
        output_path=args.output_path,
        sheet_name=args.sheet,
        data_start_row=args.start_row,
        data_end_row=args.end_row,
    )

    # 打印第一条样例
    if result["fields"]:
        print("\n" + "=" * 60)
        print("样例（第一个字段的第一条逻辑）:")
        print(json.dumps(result["fields"][0]["logics"][0],
                         ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
